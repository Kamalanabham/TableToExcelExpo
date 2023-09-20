import os
import cv2 as cv
import sys
import numpy as np
import pytesseract
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = Workbook()
ws = wb.active
ws.title="Data"



def show_wait_destroy(winname, img):   
    #return     #Displaying function
    cv.imshow(winname, img)
    cv.waitKey(0)
    cv.destroyWindow(winname)
"""
def show_wait_destroy(winname, img):        #Displaying function
    cv.imshow(winname, img)
    cv.waitKey(0)
    #cv.destroyWindow(winname)
"""
# def cropImg(i,j):
#     x1=i
#     y1=j
#     x2=x1+10
#     y2=y1+10
#     if x1>width or x2>width or y1>height or y2>height:
#         return 
#     while similarity[x1][y2]<255:
#         y2+=1
#     while similarity[x2][y1]<255:
#         x2+=1
#     cropped=thr1[x1:x2,y1:y2]
#     text = pytesseract.image_to_string(cropped,config="--psm 4")
#     file.write(text)
#     file.write("\n")
#     show_wait_destroy("Cropped", cropped)

def cropImg(i,j,h,w):
    x1=i
    y1=j
    x2=x1+10
    y2=y1+10
    if x2>=h or y2>=w:
        return 
    while similarity[x1][y2]<255:
        y2+=1
        if y2>=w:
            return
    while similarity[x2][y1]<255:
        x2+=1
        if x2>=h:
            return
    cropped=masked[x1:x2,y1:y2]
    text = pytesseract.image_to_string(cropped,config="--psm 7")
    columndata.append(text)
    file.write(text)
    file.write("\n")
    show_wait_destroy("Cropped", cropped)

#img=cv.imread('Photos/conv_example(1).png')
img=cv.imread(sys.argv[1])               #Loading image

show_wait_destroy("Original", img)                  #Loading image
print("Loading image.....")
# if len(img.shape) != 2:                     #Converting to black and white if not already.... 
#         gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
# else:                                       #img.shape returns (height , width, color_channels)
#     gray = img
gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
show_wait_destroy("BandW", gray)
print("Black and white image generated")

thr1 = cv.threshold(gray, 0, 255, cv.THRESH_BINARY_INV + cv.THRESH_OTSU)[1]
show_wait_destroy("Threshold", thr1)
print("Threshold image generated")

gray = cv.bitwise_not(gray)
bin = cv.adaptiveThreshold(gray, 255, cv.ADAPTIVE_THRESH_MEAN_C, cv.THRESH_BINARY, 15, -2)
show_wait_destroy("Binary", bin)
print("Binary image generated")

horizontal = np.copy(bin)
cols = horizontal.shape[1]              #identifies horizontal lines of table
horizontal_size = cols // 30
horizontalStructure = cv.getStructuringElement(cv.MORPH_RECT, (horizontal_size, 1))
horizontal = cv.erode(horizontal, horizontalStructure)
horizontal = cv.dilate(horizontal, horizontalStructure)
show_wait_destroy("horizontal", horizontal)
print("Horizontal lines detected")

vertical = np.copy(bin)
rows = vertical.shape[0]                #identifies vertical lines of table
vertical_size = rows // 15
verticalStructure = cv.getStructuringElement(cv.MORPH_RECT, (1,vertical_size))
vertical = cv.erode(vertical, verticalStructure)
vertical = cv.dilate(vertical, verticalStructure)
show_wait_destroy("vertical", vertical)
print("Vertical lines detected")

masked = np.copy(thr1)
for i in range(len(horizontal)):
    for j in range(len(horizontal[i])):     #Masking the table so that only the text remains
        if horizontal[i][j]==thr1[i][j] or vertical[i][j]==thr1[i][j]:
            masked[i][j]=0
show_wait_destroy("Masked", masked)
print("Image masked")

similarity = np.zeros([img.shape[0], img.shape[1], 1], dtype=np.uint8)
for i in range(len(horizontal)):
    for j in range(len(horizontal[i])):     #identifies the intersection of horizontal and vertical lines and 
        if horizontal[i][j]==vertical[i][j]:
            similarity[i][j]=horizontal[i][j]
show_wait_destroy("Similarity", similarity)
print("Coordinates identified")


height=similarity.shape[0]
width=similarity.shape[1]
print(height,width)
file = open("recognized.txt", "a")
count=0
columndata = []
for i in range(0,height-14,3):
    if(count>0):
        ws.append(columndata)
        columndata.clear()
        count=0
    for j in range(0,width-14,3):   
        if j>width:
            break  
        if similarity[i][j]==255:
            count+=1
            cropImg(i,j,height,width)
file.close
wb.save("Final.xlsx")
print("Excel sheet generated")
os.system("start soffice.exe --calc Final.xlsx")
#cv.waitKey(0)

